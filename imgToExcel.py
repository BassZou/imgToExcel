from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
from PIL import Image

workbook = Workbook()
worksheet = workbook.active
im = Image.open("test.png")
im_width = im.size[0]
im_height = im.size[1]
pix = im.load()
for row in range(1, im_height):
    for col in range(1, im_width):
        cell = worksheet.cell(column=col, row=row)
        pixpoint = pix[col - 1, row - 1]
        pixColor = "FF%02X%02X%02X" % (pixpoint[0], pixpoint[1], pixpoint[2])
        fill = PatternFill(patternType='solid', fgColor=Color(rgb=pixColor))
        cell.fill = fill
    worksheet.row_dimensions[row].height = 6
for col in range(1, im_width):
    worksheet.column_dimensions[get_column_letter(col)].width = 1
workbook.save(filename='test.xlsx')
